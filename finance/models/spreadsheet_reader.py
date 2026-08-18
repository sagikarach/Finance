"""Read an expenses file into CSV text so the existing CSV parser can handle it.

Lets the import flow accept Excel (.xlsx / .xls) and HTML exports as well as .csv:
each format is flattened into the same comma-separated shape the parser already
understands, so all formats share one code path.

Format is detected by *content*, not the file extension — Israeli banks and
credit-card issuers routinely hand out an HTML table saved as ``.xls`` (Bank
Leumi's "Excel" export is the classic case), or an .xlsx renamed to .xls. We
sniff the leading bytes and dispatch accordingly, so a mislabeled file still
loads. openpyxl (.xlsx) / xlrd (real BIFF .xls) are optional deps imported
lazily; the HTML path uses only the standard library.
"""

from __future__ import annotations

import csv
from datetime import date, datetime
from html.parser import HTMLParser
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Iterable, List, Sequence


def _cell_to_str(value: Any) -> str:
    """Render a spreadsheet cell the way the CSV parser expects: dates as
    dd/mm/yyyy, whole floats without a trailing .0, everything else as text."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _rows_to_csv_text(rows: Iterable[Sequence[Any]]) -> str:
    buf = StringIO()
    writer = csv.writer(buf)
    for row in rows:
        writer.writerow([_cell_to_str(c) for c in row])
    return buf.getvalue()


class _TableExtractor(HTMLParser):
    """Collect every <tr>'s cells across all <table>s in the document.

    Multiple tables/sections are simply concatenated: the downstream parser
    scans for its own header row and ignores everything else, so extra rows
    (titles, totals, a second "future charges" section) are harmless.
    """

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.rows: List[List[str]] = []
        self._row: List[str] | None = None
        self._cell: List[str] | None = None
        self._in_cell = False

    def handle_starttag(self, tag: str, attrs: Any) -> None:
        if tag == "tr":
            self._row = []
        elif tag in ("td", "th"):
            self._cell = []
            self._in_cell = True
        elif tag == "br" and self._in_cell and self._cell is not None:
            self._cell.append(" ")

    def handle_endtag(self, tag: str) -> None:
        if tag == "tr" and self._row is not None:
            if any(c.strip() for c in self._row):
                self.rows.append(self._row)
            self._row = None
        elif tag in ("td", "th") and self._cell is not None:
            text = " ".join("".join(self._cell).split())
            if self._row is not None:
                self._row.append(text)
            self._cell = None
            self._in_cell = False

    def handle_data(self, data: str) -> None:
        if self._in_cell and self._cell is not None:
            self._cell.append(data)


def _html_to_csv_text(text: str) -> str:
    extractor = _TableExtractor()
    extractor.feed(text)
    return _rows_to_csv_text(extractor.rows)


def _xlsx_to_csv_text(data: bytes) -> str:
    from openpyxl import load_workbook  # optional dep

    # Load from bytes, not the path, so openpyxl dispatches on the ZIP content
    # rather than refusing a .xlsx that happens to carry a .xls extension.
    wb = load_workbook(filename=BytesIO(data), read_only=True, data_only=True)
    try:
        ws = wb.active
        if ws is None:
            return ""
        return _rows_to_csv_text(ws.iter_rows(values_only=True))
    finally:
        wb.close()


def _xls_to_csv_text(data: bytes) -> str:
    import xlrd  # optional dep

    # From bytes (Drive downloads, temp-less), not a path.
    book = xlrd.open_workbook(file_contents=data)
    sheet = book.sheet_by_index(0)
    rows: list[list[Any]] = []
    for r in range(sheet.nrows):
        row: list[Any] = []
        for c in range(sheet.ncols):
            cell = sheet.cell(r, c)
            if cell.ctype == xlrd.XL_CELL_DATE:
                y, mo, d, *_ = xlrd.xldate_as_tuple(float(cell.value), book.datemode)
                row.append(f"{d:02d}/{mo:02d}/{y:04d}")
            else:
                row.append(cell.value)
        rows.append(row)
    return _rows_to_csv_text(rows)


def _looks_like_html(head: bytes) -> bool:
    lowered = head.lstrip().lower()
    return (
        lowered.startswith(b"<!doctype html")
        or lowered.startswith(b"<html")
        or lowered.startswith(b"<?xml")
        or b"<table" in lowered
        or b"<html" in lowered
    )


def bytes_to_csv_text(raw: bytes, *, suffix: str = "") -> str:
    """CSV text for any supported expenses file already read into memory.

    Dispatch is by content signature first (so a mislabeled file still loads),
    falling back to *suffix*:

    * ZIP magic ``PK\\x03\\x04``            -> .xlsx via openpyxl
    * OLE2 magic ``\\xd0\\xcf\\x11\\xe0``   -> real .xls via xlrd
    * looks like HTML                       -> HTML table(s) (stdlib)
    * otherwise                             -> plain text, tolerant of a UTF-8 BOM

    Shared by the local file importer and the Drive-inbox importer so both go
    through exactly the same format handling.
    """
    head = raw[:512]

    if head.startswith(b"PK\x03\x04"):
        return _xlsx_to_csv_text(raw)
    if head.startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"):
        return _xls_to_csv_text(raw)
    if _looks_like_html(head):
        return _html_to_csv_text(raw.decode("utf-8", errors="replace"))

    # Fall back to the extension for anything the signature didn't catch.
    s = str(suffix or "").lower()
    if s == ".xlsx":
        return _xlsx_to_csv_text(raw)
    if s == ".xls":
        return _xls_to_csv_text(raw)

    # .csv / .txt / unknown → plain text, tolerant of a UTF-8 BOM.
    return raw.decode("utf-8-sig", errors="replace")


def file_to_csv_text(path: Path) -> str:
    """CSV text for any supported expenses file on disk (see :func:`bytes_to_csv_text`)."""
    return bytes_to_csv_text(path.read_bytes(), suffix=path.suffix)
